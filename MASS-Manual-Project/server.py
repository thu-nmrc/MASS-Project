# MASS Project - originated by Xiaoli Hu (胡晓李)
# Copyright (c) 2025 Xiaoli Hu
# Licensed under the MIT License. See LICENSE file in the project root.
# MASS Server - Multi-Agent Social Simulation Platform
# Version: 1.0.0

import http.server
import socketserver
import json
import urllib.request
import urllib.error
import os
import sys
import time
import random
import socket
import webbrowser
import threading
import csv
from datetime import datetime

# 尝试导入 openpyxl，如果没有则提示安装
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARNING] openpyxl 未安装，Excel 导出功能不可用")
    print("[INFO] 安装方法: pip install openpyxl")

# Configuration
PORT = int(os.environ.get('MASS_PORT', '8899'))

def _resolve_root():
    """
    解析项目根目录：
    - 若设置了 MASS_BASE_DIR 环境变量且存在，则使用它
    - 打包(frozen)时，优先 cwd，其次 exe 所在目录，或向上2层寻找含 exports 的目录
    - 普通脚本运行时，优先 cwd，其次脚本目录，或其父目录中包含 exports 的目录
    """
    env_base = os.environ.get("MASS_BASE_DIR")
    if env_base and os.path.isdir(env_base):
        return os.path.abspath(env_base)

    try:
        if getattr(sys, 'frozen', False):
            argv0 = os.path.abspath(sys.argv[0]) if sys.argv and sys.argv[0] else ""
            argv0_dir = os.path.dirname(argv0) if argv0 else ""

            exe_dir = os.path.dirname(os.path.abspath(sys.executable))
            mei_dir = os.path.abspath(getattr(sys, "_MEIPASS", "")) if getattr(sys, "_MEIPASS", None) else ""

            if argv0_dir and (not mei_dir or os.path.normcase(argv0_dir) != os.path.normcase(mei_dir)):
                return argv0_dir

            if exe_dir and (not mei_dir or os.path.normcase(exe_dir) != os.path.normcase(mei_dir)):
                return exe_dir

            cwd = os.path.abspath(os.getcwd())
            if os.path.isdir(os.path.join(cwd, 'exports')):
                return cwd

            base_candidates = []
            if argv0_dir:
                base_candidates.append(argv0_dir)
            base_candidates.append(exe_dir)

            for base in base_candidates:
                if mei_dir and os.path.normcase(base) == os.path.normcase(mei_dir):
                    continue
                return base

            return cwd
        else:
            # 普通脚本
            script_dir = os.path.dirname(os.path.abspath(__file__))
            for candidate in [
                os.getcwd(),
                script_dir,
                os.path.abspath(os.path.join(script_dir, '..')),
            ]:
                if os.path.isdir(os.path.join(candidate, 'exports')):
                    return candidate
            return script_dir
    except Exception:
        return os.getcwd()

ROOT = _resolve_root()
STATIC_DIR = os.path.join(ROOT, 'static')
EXPORT_DIR = os.path.join(ROOT, 'exports')
CSV_FILE = os.path.join(EXPORT_DIR, 'messages.csv')

# Ensure export directory exists
os.makedirs(EXPORT_DIR, exist_ok=True)

# Global configuration
CONFIG = {
    "baseUrl": "",
    "apiKey": "",
    "modelName": "",
    "maxTokens": 512,
    "multiApi": False
}

RETRY_CODES = {429, 500, 502, 503, 504}


def init_csv():
    """Initialize CSV file with headers"""
    headers = ['Agent名称', '系统消息', '用户消息', 'AI回复', '回合', '时间戳', 'Agent ID', 'Tokens', '延迟(ms)']
    with open(CSV_FILE, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
    print(f"[MASS] CSV initialized: {CSV_FILE}")


def save_to_csv(round_num, agent_id, agent_name, system_msg, user_msg, response, tokens=0, latency=0):
    """Save message to CSV"""
    try:
        if not os.path.exists(CSV_FILE):
            init_csv()
        
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        data = [agent_name, system_msg, user_msg, response, round_num, timestamp, agent_id, tokens, latency]
        
        with open(CSV_FILE, 'a', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(data)
        
        print(f"[MASS] Saved: Agent {agent_id}, Round {round_num}")
    except Exception as e:
        print(f"[MASS] CSV save error: {e}")


def export_csv():
    """Export CSV with sequence number"""
    try:
        if not os.path.exists(CSV_FILE):
            return {"ok": False, "error": "No messages file found"}
        
        seq = 1
        while True:
            filename = f"export_{seq}.csv"
            path = os.path.join(EXPORT_DIR, filename)
            if not os.path.exists(path):
                break
            seq += 1
        
        headers = ['Agent名称', '发送的Prompt', 'AI回复', '回合']
        
        with open(CSV_FILE, 'r', encoding='utf-8-sig') as src:
            reader = csv.reader(src)
            next(reader)  # Skip header
            
            with open(path, 'w', newline='', encoding='utf-8-sig') as dst:
                writer = csv.writer(dst)
                writer.writerow(headers)
                
                for row in reader:
                    if len(row) >= 9:
                        agent_name, sys_msg, user_msg, response, round_num = row[0], row[1], row[2], row[3], row[4]
                        prompt = f"系统: {sys_msg}\n用户: {user_msg}" if sys_msg else f"用户: {user_msg}"
                        writer.writerow([agent_name, prompt.strip(), response, round_num])
        
        return {"ok": True, "filename": filename, "path": path}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def export_excel():
    """Export Excel with sequence number"""
    if not EXCEL_AVAILABLE:
        return {"ok": False, "error": "openpyxl 未安装，请运行: pip install openpyxl"}
    
    try:
        if not os.path.exists(CSV_FILE):
            return {"ok": False, "error": "No messages file found"}
        
        # 找到下一个可用的序号
        seq = 1
        while True:
            filename = f"export_{seq}.xlsx"
            path = os.path.join(EXPORT_DIR, filename)
            if not os.path.exists(path):
                break
            seq += 1
        
        # 创建 Excel 工作簿
        wb = Workbook()
        
        # 创建"模拟结果"工作表
        ws_summary = wb.active
        ws_summary.title = "模拟结果"
        
        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入表头
        headers = ['Agent名称', '发送的Prompt', 'AI回复', '回合']
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 读取 CSV 数据并写入 Excel
        with open(CSV_FILE, 'r', encoding='utf-8-sig') as src:
            reader = csv.reader(src)
            next(reader)  # 跳过表头
            
            row_num = 2
            for row in reader:
                if len(row) >= 9:
                    agent_name = row[0]
                    sys_msg = row[1] or ""
                    user_msg = row[2] or ""
                    response = row[3]
                    round_num = row[4]
                    
                    # 合并系统消息和用户消息
                    prompt = f"系统: {sys_msg}\n用户: {user_msg}" if sys_msg else f"用户: {user_msg}"
                    
                    # 写入数据
                    ws_summary.cell(row=row_num, column=1, value=agent_name)
                    ws_summary.cell(row=row_num, column=2, value=prompt.strip())
                    ws_summary.cell(row=row_num, column=3, value=response)
                    ws_summary.cell(row=row_num, column=4, value=int(round_num) if round_num.isdigit() else round_num)
                    
                    row_num += 1
        
        # 创建"详细数据"工作表
        ws_detail = wb.create_sheet("详细数据")
        
        # 详细数据表头
        detail_headers = ['Agent名称', '系统消息', '用户消息', 'AI回复', '回合', '时间戳', 'Agent ID', 'Tokens', '延迟(ms)']
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 写入详细数据
        with open(CSV_FILE, 'r', encoding='utf-8-sig') as src:
            reader = csv.reader(src)
            next(reader)  # 跳过表头
            
            for row_num, row in enumerate(reader, 2):
                for col_num, value in enumerate(row, 1):
                    ws_detail.cell(row=row_num, column=col_num, value=value)
        
        # 调整列宽
        for ws in [ws_summary, ws_detail]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存 Excel 文件
        wb.save(path)
        
        return {"ok": True, "filename": filename, "path": path}
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"ok": False, "error": str(e)}


def json_bytes(obj):
    return json.dumps(obj, ensure_ascii=False).encode('utf-8')


def post_with_retry(url, data, headers, timeout=60, retries=3):
    """POST request with retry logic"""
    for attempt in range(1, retries + 1):
        try:
            req = urllib.request.Request(url, data=data, headers=headers, method="POST")
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                code = resp.getcode()
                text = resp.read().decode('utf-8', 'ignore')
                if 200 <= code < 300:
                    return {"ok": True, "status": code, "text": text, "attempts": attempt}
                if code not in RETRY_CODES:
                    return {"ok": False, "status": code, "text": text, "attempts": attempt}
        except urllib.error.HTTPError as e:
            code = e.code
            text = e.read().decode('utf-8', 'ignore') if hasattr(e, 'read') else str(e)
            if code not in RETRY_CODES or attempt == retries:
                return {"ok": False, "status": code, "text": text, "attempts": attempt}
        except Exception as e:
            if attempt == retries:
                return {"ok": False, "status": 0, "text": str(e), "attempts": attempt}
        
        time.sleep((1.5 ** attempt) + random.uniform(0, 0.3))
    
    return {"ok": False, "status": 0, "text": "Max retries exceeded", "attempts": retries}


def call_llm(cfg, body):
    """Call LLM API"""
    base_url = (body.get('baseUrl') or cfg.get('baseUrl') or '').rstrip('/')
    if not base_url:
        print("[ERROR] Missing baseUrl")
        return {"status": 400, "error": "Missing baseUrl", "head": "Missing baseUrl"}
    
    url = base_url + "/chat/completions"
    model = body.get('model') or cfg.get('modelName') or ''
    
    if not model:
        print("[ERROR] Missing model name")
        return {"status": 400, "error": "Missing model name", "head": "Missing model name"}
    
    messages = body.get('messages') or []
    
    if not messages:
        print("[ERROR] Missing messages")
        return {"status": 400, "error": "Missing messages", "head": "Missing messages"}
    
    headers = {"Content-Type": "application/json"}
    api_key = body.get('apiKey') or cfg.get('apiKey') or ''
    if api_key:
        headers["Authorization"] = f"Bearer {api_key}"
    
    payload = {"model": model, "messages": messages}
    
    print(f"[DEBUG] Calling API: {url}")
    print(f"[DEBUG] Model: {model}")
    print(f"[DEBUG] Messages count: {len(messages)}")
    
    res = post_with_retry(url, json_bytes(payload), headers)
    
    print(f"[DEBUG] API response status: {res.get('status')}")
    print(f"[DEBUG] API response ok: {res.get('ok')}")
    
    try:
        parsed = json.loads(res.get("text", "{}"))
    except Exception as e:
        print(f"[ERROR] Failed to parse JSON response: {e}")
        parsed = None
    
    status_msg = ""
    if isinstance(parsed, dict):
        if "error" in parsed:
            err = parsed.get("error", {})
            status_msg = err.get("message") or err.get("type") or "API Error"
            print(f"[ERROR] API returned error: {status_msg}")
        elif "id" in parsed:
            status_msg = "ok"
            print("[DEBUG] API call successful")
        else:
            status_msg = "Unknown response format"
            print(f"[WARNING] Unknown response format: {parsed}")
    else:
        status_msg = "Invalid JSON response"
        print(f"[ERROR] Invalid JSON response")
    
    return {
        "status": res.get("status", 0),
        "json": parsed,
        "raw": res.get("text", ""),
        "attempts": res.get("attempts", 1),
        "head": status_msg,
        "ok": res.get("ok", False)
    }


class RequestHandler(http.server.SimpleHTTPRequestHandler):
    def _headers(self, code=200, ctype="application/json; charset=utf-8"):
        self.send_response(code)
        self.send_header("Content-Type", ctype)
        self.send_header("Cache-Control", "no-cache")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
    
    def translate_path(self, path):
        if path in ["/", ""]:
            return os.path.join(STATIC_DIR, "index.html")
        if path.startswith("/static/"):
            return os.path.join(ROOT, path.lstrip("/"))
        return os.path.join(STATIC_DIR, "index.html")
    
    def do_GET(self):
        if self.path.startswith("/api/"):
            self._headers(404)
            self.wfile.write(json_bytes({"status": 404, "error": "Not Found"}))
            return
        return http.server.SimpleHTTPRequestHandler.do_GET(self)
    
    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", "0") or "0")
            body = self.rfile.read(length).decode("utf-8", "ignore") if length else "{}"
            try:
                payload = json.loads(body) if body else {}
            except:
                payload = {}
            
            # Config endpoint
            if self.path == "/api/config":
                for k in ["baseUrl", "apiKey", "modelName", "multiApi"]:
                    if k in payload:
                        CONFIG[k] = payload[k]
                self._headers(200)
                self.wfile.write(json_bytes({"status": 200, "head": "saved", "config": {**CONFIG, "apiKey": "***"}}))
                return
            
            # Test endpoint
            if self.path == "/api/test":
                try:
                    # 打印调试信息
                    print(f"[DEBUG] Test API request payload: {payload}")
                    
                    cfg = {**CONFIG}
                    # 从 payload 中获取参数
                    if 'baseUrl' in payload:
                        cfg['baseUrl'] = payload['baseUrl']
                    if 'apiKey' in payload:
                        cfg['apiKey'] = payload['apiKey']
                    if 'model' in payload:
                        cfg['modelName'] = payload['model']
                    
                    messages = payload.get("messages") or [{"role": "system", "content": "ping"}, {"role": "user", "content": "ping"}]
                    
                    print(f"[DEBUG] Calling LLM with config: baseUrl={cfg.get('baseUrl')}, model={cfg.get('modelName')}")
                    
                    res = call_llm(cfg, {**payload, "messages": messages})
                    
                    print(f"[DEBUG] LLM response: status={res.get('status')}, head={res.get('head')}")
                    
                    self._headers(200)
                    self.wfile.write(json_bytes(res))
                except Exception as e:
                    print(f"[ERROR] Test API error: {e}")
                    import traceback
                    traceback.print_exc()
                    self._headers(500)
                    self.wfile.write(json_bytes({"status": 500, "error": str(e), "head": "Server Error"}))
                return
            
            # Complete endpoint
            if self.path == "/api/complete":
                res = call_llm(CONFIG, payload)
                self._headers(200)
                self.wfile.write(json_bytes(res))
                return
            
            # Save message endpoint
            if self.path == "/api/save_message":
                save_to_csv(
                    payload.get("round", 1),
                    payload.get("agentId", ""),
                    payload.get("agentName", ""),
                    payload.get("systemText", ""),
                    payload.get("userText", ""),
                    payload.get("responseText", ""),
                    payload.get("tokens", 0),
                    payload.get("latency", 0)
                )
                self._headers(200)
                self.wfile.write(json_bytes({"status": 200, "message": "Saved"}))
                return
            
            # Export CSV endpoint
            if self.path == "/api/export":
                result = export_csv()
                if result["ok"]:
                    self._headers(200)
                    self.wfile.write(json_bytes({
                        "status": 200,
                        "filename": result["filename"],
                        "path": result.get("path", ""),
                        "export_dir": EXPORT_DIR,
                        "root": ROOT,
                        "message": f"Exported as {result['filename']}"
                    }))
                else:
                    self._headers(500)
                    self.wfile.write(json_bytes({"status": 500, "error": result["error"]}))
                return
            
            # Export Excel endpoint
            if self.path == "/api/export_excel":
                result = export_excel()
                if result["ok"]:
                    self._headers(200)
                    self.wfile.write(json_bytes({
                        "status": 200,
                        "filename": result["filename"],
                        "path": result.get("path", ""),
                        "export_dir": EXPORT_DIR,
                        "root": ROOT,
                        "message": f"Exported as {result['filename']}",
                        "excel_available": True
                    }))
                else:
                    self._headers(500)
                    self.wfile.write(json_bytes({"status": 500, "error": result["error"], "excel_available": EXCEL_AVAILABLE}))
                return
            
            # Check Excel availability endpoint
            if self.path == "/api/check_excel":
                self._headers(200)
                self.wfile.write(json_bytes({"status": 200, "excel_available": EXCEL_AVAILABLE}))
                return
            
            # Init CSV endpoint
            if self.path == "/api/init_csv":
                init_csv()
                self._headers(200)
                self.wfile.write(json_bytes({"status": 200, "message": "CSV initialized"}))
                return
            
            self._headers(404)
            self.wfile.write(json_bytes({"status": 404, "error": "Unknown endpoint"}))
        except Exception as e:
            self._headers(500)
            self.wfile.write(json_bytes({"status": 500, "error": str(e)}))


def open_browser():
    """Open browser after short delay"""
    def task():
        try:
            time.sleep(0.5)
            webbrowser.open(f"http://127.0.0.1:{PORT}/", new=2)
        except:
            pass
    threading.Thread(target=task, daemon=True).start()


def run():
    """Start server"""
    os.chdir(ROOT)
    init_csv()
    
    with socketserver.TCPServer(("127.0.0.1", PORT), RequestHandler) as httpd:
        print(f"[MASS] Server running at http://127.0.0.1:{PORT}")
        print(f"[MASS] Press Ctrl+C to stop")
        
        # if not os.environ.get("MASS_NO_OPEN"):
        #     open_browser()
        #
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print("\n[MASS] Server stopped")


if __name__ == "__main__":
    run()
